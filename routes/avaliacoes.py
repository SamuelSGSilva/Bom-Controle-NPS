from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from database import get_db
from models.models import Avaliacao, Cliente
from pydantic import BaseModel
from auth import verificar_token
import openpyxl
from io import BytesIO
from services.google_sheets_sync import buscar_registros_google_sheets, converter_linha_para_payload
from services.auditoria import registrar_log

router = APIRouter()

def _nome_cliente(db: Session, cliente_id: int) -> str | None:
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    return cliente.nome if cliente else None

def _serializar_avaliacao(a: Avaliacao, cliente_nome: str | None):
    dados = {c.name: getattr(a, c.name) for c in Avaliacao.__table__.columns}
    dados["cliente_nome"] = cliente_nome
    return dados

class AvaliacaoSchema(BaseModel):
    cliente_id: int
    numero_os: str | None = None
    produto_servico: str | None = None
    data_atendimento: datetime | None = None
    cidade_estado: str | None = None
    nota_primeiro_contato: int | None = None
    nota_clareza_informacoes: int | None = None
    nota_processo_fechamento: int | None = None
    nota_link_pagamento: int | None = None
    nota_nota_fiscal: int | None = None
    nota_entrega_prazo: int | None = None
    nota_embalagem: int | None = None
    nota_entrega_tecnica: int | None = None
    nota_suporte_produto: int | None = None
    avaliacao_geral: int | None = None
    nps_score: int | None = None
    o_que_gostou: str | None = None
    o_que_melhorar: str | None = None
    cs_responsavel: str | None = None

@router.get("/avaliacoes")
def listar_avaliacoes(db: Session = Depends(get_db), token: dict = Depends(verificar_token)):
    return db.query(Avaliacao).all()

@router.post("/avaliacoes/publico")
def criar_avaliacao_publico(avaliacao: AvaliacaoSchema, db: Session = Depends(get_db)):
    nova = Avaliacao(**avaliacao.model_dump())
    db.add(nova)
    db.flush()
    dados_depois = _serializar_avaliacao(nova, _nome_cliente(db, nova.cliente_id))
    registrar_log(db, None, "create", "avaliacao", nova.id, None, dados_depois)
    db.commit()
    db.refresh(nova)
    return {"message": "Avaliação registrada com sucesso!"}

@router.post("/avaliacoes")
def criar_avaliacao(avaliacao: AvaliacaoSchema, db: Session = Depends(get_db), token: dict = Depends(verificar_token)):
    nova = Avaliacao(**avaliacao.model_dump())
    db.add(nova)
    db.flush()
    dados_depois = _serializar_avaliacao(nova, _nome_cliente(db, nova.cliente_id))
    registrar_log(db, token.get("sub"), "create", "avaliacao", nova.id, None, dados_depois)
    db.commit()
    db.refresh(nova)
    return nova

@router.delete("/avaliacoes/{id}")
def deletar_avaliacao(id: int, db: Session = Depends(get_db), token: dict = Depends(verificar_token)):
    avaliacao = db.query(Avaliacao).filter(Avaliacao.id == id).first()
    if not avaliacao:
        raise HTTPException(status_code=404, detail="Avaliação não encontrada")

    dados_antes = _serializar_avaliacao(avaliacao, _nome_cliente(db, avaliacao.cliente_id))
    registrar_log(db, token.get("sub"), "delete", "avaliacao", avaliacao.id, dados_antes, None)
    db.delete(avaliacao)
    db.commit()
    return {"message": "Avaliação deletada com sucesso"}

@router.get("/avaliacoes/relatorio/nps")
def relatorio_nps(db: Session = Depends(get_db), token: dict = Depends(verificar_token)):
    todas = db.query(Avaliacao).all()
    total = len(todas)

    if total == 0:
        return {"message": "Nenhuma avaliação encontrada"}

    com_nps = [a for a in todas if a.nps_score is not None]
    promotores = len([a for a in com_nps if a.nps_score >= 9])
    neutros = len([a for a in com_nps if 7 <= a.nps_score <= 8])
    detratores = len([a for a in com_nps if a.nps_score <= 6])

    nps = ((promotores - detratores) / len(com_nps)) * 100 if com_nps else 0

    return {
        "total_avaliacoes": total,
        "promotores": promotores,
        "neutros": neutros,
        "detratores": detratores,
        "nps": round(nps, 2)
    }

@router.get("/avaliacoes/{cliente_id}")
def buscar_avaliacoes_cliente(cliente_id: int, db: Session = Depends(get_db), token: dict = Depends(verificar_token)):
    return db.query(Avaliacao).filter(Avaliacao.cliente_id == cliente_id).all()

@router.get("/avaliacoes/relatorio/painel")
def painel_resultados(db: Session = Depends(get_db), token: dict = Depends(verificar_token)):
    avaliacoes = db.query(Avaliacao).all()
    total = len(avaliacoes)

    if total == 0:
        return {"message": "Nenhuma avaliação encontrada"}

    # Alguns campos usam valores acima de 5 pra opções que não são nota
    # (nota_link_pagamento: 6=Pix, 7=Boleto; nota_entrega_prazo/nota_embalagem: 6=Não tem).
    # Esses valores precisam ficar fora das médias, senão inflam a nota real.
    VALORES_NAO_NUMERICOS = {
        'nota_link_pagamento': {6, 7},
        'nota_entrega_prazo': {6},
        'nota_embalagem': {6},
    }

    def eh_nota_valida(campo, valor):
        return valor is not None and valor not in VALORES_NAO_NUMERICOS.get(campo, set())

    def media(campo):
        valores = [getattr(a, campo) for a in avaliacoes if eh_nota_valida(campo, getattr(a, campo))]
        return round(sum(valores) / len(valores), 2) if valores else None

    def status_nps(score):
        if score >= 75: return "Excelência"
        if score >= 50: return "Qualidade"
        if score >= 0: return "Aperfeiçoamento"
        return "Ponto de atenção"

    notas_todas = []
    campos_notas = [
        'nota_primeiro_contato', 'nota_clareza_informacoes', 'nota_processo_fechamento',
        'nota_link_pagamento', 'nota_nota_fiscal', 'nota_entrega_prazo',
        'nota_embalagem', 'nota_entrega_tecnica', 'nota_suporte_produto', 'avaliacao_geral'
    ]
    for a in avaliacoes:
        for campo in campos_notas:
            v = getattr(a, campo)
            if eh_nota_valida(campo, v):
                notas_todas.append(v)

    media_geral = round(sum(notas_todas) / len(notas_todas), 2) if notas_todas else None
    nota_min = min(notas_todas) if notas_todas else None
    nota_max = max(notas_todas) if notas_todas else None

    nps_scores = [a.nps_score for a in avaliacoes if a.nps_score is not None]
    promotores = len([n for n in nps_scores if n >= 9])
    detratores = len([n for n in nps_scores if n <= 6])
    nps = round(((promotores - detratores) / len(nps_scores)) * 100, 2) if nps_scores else 0

    return {
        "resumo": {
            "total_avaliacoes": total,
            "media_geral": media_geral,
            "nota_min": nota_min,
            "nota_max": nota_max
        },
        "nps": {
            "score": nps,
            "status": status_nps(nps)
        },
        "medias_por_bloco": {
            "vendas_primeiro_contato": media('nota_primeiro_contato'),
            "vendas_clareza": media('nota_clareza_informacoes'),
            "vendas_fechamento": media('nota_processo_fechamento'),
            "financeiro_link": media('nota_link_pagamento'),
            "financeiro_nf": media('nota_nota_fiscal'),
            "transporte_prazo": media('nota_entrega_prazo'),
            "transporte_embalagem": media('nota_embalagem'),
            "suporte_entrega": media('nota_entrega_tecnica'),
            "suporte_produto": media('nota_suporte_produto'),
            "avaliacao_geral": media('avaliacao_geral')
        }
    }

@router.get("/avaliacoes/relatorio/evolucao")
def evolucao_nps(db: Session = Depends(get_db), token: dict = Depends(verificar_token)):
    avaliacoes = db.query(Avaliacao).filter(Avaliacao.nps_score.isnot(None)).all()

    if not avaliacoes:
        return {"message": "Nenhuma avaliação encontrada"}

    grupos = {}
    for a in avaliacoes:
        data_ref = a.data_atendimento or a.created_at
        if not data_ref:
            continue
        chave = data_ref.strftime("%Y-%m")
        grupos.setdefault(chave, []).append(a.nps_score)

    evolucao = []
    for chave in sorted(grupos.keys()):
        scores = grupos[chave]
        total = len(scores)
        promotores = len([s for s in scores if s >= 9])
        detratores = len([s for s in scores if s <= 6])
        nps = round(((promotores - detratores) / total) * 100, 2)
        evolucao.append({
            "mes": chave,
            "total_avaliacoes": total,
            "nps": nps
        })

    return {"evolucao": evolucao}


@router.post("/avaliacoes/importar")
async def importar_planilha(file: UploadFile = File(...), db: Session = Depends(get_db), token: dict = Depends(verificar_token)):
    conteudo = await file.read()
    wb = openpyxl.load_workbook(BytesIO(conteudo), data_only=True)

    if "Registros" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Aba 'Registros' não encontrada na planilha")

    ws = wb["Registros"]

    importados = 0

    for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        nome_cliente = row[2]

        if not nome_cliente:
            continue

        cliente = db.query(Cliente).filter(Cliente.nome == nome_cliente).first()

        if not cliente:
            cliente = Cliente(
                nome=nome_cliente,
                email=f"sem-email-{i}@autoluiz.com"
            )
            db.add(cliente)
            db.commit()
            db.refresh(cliente)

        payload, _ = converter_linha_para_payload(list(row), i)
        if not payload:
            continue

        nova = Avaliacao(cliente_id=cliente.id, **payload)
        db.add(nova)
        importados += 1

    db.commit()
    return {"message": f"{importados} avaliações importadas com sucesso!"}


@router.post("/avaliacoes/importar-google-sheets")
def importar_google_sheets(db: Session = Depends(get_db), token: dict = Depends(verificar_token)):
    try:
        rows = buscar_registros_google_sheets()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    importados = 0

    for i, row in enumerate(rows[3:], start=4):
        payload, nome_cliente = converter_linha_para_payload(row, i)
        if not payload or not nome_cliente:
            continue

        cliente = db.query(Cliente).filter(Cliente.nome == nome_cliente).first()
        if not cliente:
            cliente = Cliente(nome=nome_cliente, email=f"sem-email-{i}@autoluiz.com")
            db.add(cliente)
            db.commit()
            db.refresh(cliente)

        nova = Avaliacao(cliente_id=cliente.id, **payload)
        db.add(nova)
        importados += 1

    db.commit()
    return {"message": f"{importados} avaliações importadas do Google Sheets com sucesso!"}